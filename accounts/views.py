import openpyxl
from openpyxl import styles

from django.conf import settings
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied
from django.core.urlresolvers import reverse_lazy
from django.db.models.functions import Lower
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import View, TemplateView
from django.views.generic.edit import FormView

from month import Month
from openpyxl.styles.named_styles import NamedStyle
from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter

from rest_framework import viewsets, permissions
from rest_framework.decorators import list_route, detail_route
from rest_framework.exceptions import ParseError
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response

from userena import settings as userena_settings
from userena.models import UserenaSignup
from userena.views import signup, profile_edit, profile_detail

from accounts.filters import InternshipFilter
from accounts.forms import ChooseUniversityForm, KSAUHSSignupForm, AGUSignupForm, OutsideSignupForm, \
    KSAUHSProfileEditForm, AGUProfileEditForm, OutsideProfileEditForm, ResendForm
from accounts.models import Profile, Intern, University, Batch
from accounts.permissions import IsStaff
from accounts.serializers import ProfileSerializer, InternSerializer, UserSerializer, InternTableSerializer, \
    BatchSerializer
from hospitals.models import Department
from months.models import Internship
from months.serializers import FullInternshipSerializer2
from rotations.models import Rotation
from rotations.serializers import FullRotationSerializer


class SignupWrapper(View):
    def get(self, request, *args, **kwargs):
        context = {'form': ChooseUniversityForm}
        return render(request, 'accounts/signup_start.html', context)

    def get_signup_form(self, university_id):
        if university_id == -1:
            form = OutsideSignupForm
        else:
            university = University.objects.get(id=university_id)
            if university.is_ksauhs:
                form = KSAUHSSignupForm
            elif university.is_agu:
                form = AGUSignupForm
            else:
                form = OutsideSignupForm
        form.university_id = university_id
        return form

    def post(self, request, *args, **kwargs):
        page = request.POST.get('page')
        if int(page) == 1:
            form = ChooseUniversityForm(request.POST)
            if form.is_valid():
                university_id = int(form.cleaned_data.get('university_id'))

                signup_form = self.get_signup_form(university_id)

                request.method = "GET"  # Return the `signup` output as if it were a GET request
                return signup(request, signup_form=signup_form)
            return render(request, 'accounts/signup_start.html', {'form': form})

        elif int(page) == 2:
            university_id = int(request.POST.get('university'))
            signup_form = self.get_signup_form(university_id)
            return signup(request, signup_form=signup_form)


class ProfileDetailWrapper(View):
    def get_template_name(self, user):
        intern_profile = user.profile.intern
        if intern_profile.is_ksauhs_intern:
            return 'accounts/profile_detail/ksauhs.html'
        elif intern_profile.is_agu_intern:
            return 'accounts/profile_detail/agu.html'
        return 'accounts/profile_detail/outside.html'

    def get(self, request, *args, **kwargs):
        user = User.objects.get_by_natural_key(kwargs.get('username'))

        if not hasattr(user.profile, 'intern'):
            raise PermissionDenied

        kwargs['template_name'] = self.get_template_name(user)
        return profile_detail(request, *args, **kwargs)


class ProfileEditWrapper(View):
    def get_profile_edit_form(self, user):
        intern_profile = user.profile.intern
        if intern_profile.is_ksauhs_intern:
            return KSAUHSProfileEditForm
        if intern_profile.is_agu_intern:
            return AGUProfileEditForm
        return OutsideProfileEditForm

    def get_template_name(self, user):
        intern_profile = user.profile.intern
        if intern_profile.is_ksauhs_intern:
            return 'accounts/profile_form/ksauhs.html'
        elif intern_profile.is_agu_intern:
            return 'accounts/profile_form/agu.html'
        return 'accounts/profile_form/outside.html'

    def get(self, request, *args, **kwargs):
        user = User.objects.get_by_natural_key(kwargs.get('username'))
        kwargs['edit_profile_form'] = self.get_profile_edit_form(user)
        kwargs['template_name'] = self.get_template_name(user)
        return profile_edit(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        user = User.objects.get_by_natural_key(kwargs.get('username'))
        kwargs['edit_profile_form'] = self.get_profile_edit_form(user)
        kwargs['template_name'] = self.get_template_name(user)
        return profile_edit(request, *args, **kwargs)


class ResendConfirmationKey(FormView):
    template_name = "accounts/resend_confirmation_code.html"
    form_class = ResendForm
    success_url = reverse_lazy('resend_activation_complete')

    def form_valid(self, form):
        """
        Issue a new activation if form is valid.
        """
        UserenaSignup.objects.reissue_activation(form.user.userena_signup.activation_key)
        return redirect(self.get_success_url())


class ResendConfirmationKeyComplete(TemplateView):
    template_name = "userena/activate_retry_success.html"

    def get_context_data(self, **kwargs):
        context = super(ResendConfirmationKeyComplete, self).get_context_data(**kwargs)
        context['userena_activation_days'] = userena_settings.USERENA_ACTIVATION_DAYS
        context['SUPPORT_EMAIL_ADDRESS'] = settings.SUPPORT_EMAIL_ADDRESS
        return context


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = UserSerializer
    queryset = User.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if self.request.user.has_perm("accounts.user.view_all"):
            return self.queryset.all()
        return self.queryset.filter(username=self.request.user.username)


class ProfileViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ProfileSerializer
    queryset = Profile.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if self.request.user.has_perm("accounts.profile.view_all"):
            return self.queryset.all()
        return self.queryset.filter(user=self.request.user)


class InternViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = InternSerializer
    queryset = Intern.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if self.request.user.has_perm("accounts.intern.view_all"):
            return self.queryset.all()
        return self.queryset.filter(profile__user=self.request.user)

    @list_route(methods=['get'], permission_classes=[permissions.IsAuthenticated, IsStaff])
    def as_table(self, request, *args, **kwargs):
        interns = self.queryset.all().prefetch_related('profile__user', 'internship')
        serialized = InternTableSerializer(interns, many=True)
        return Response(serialized.data)


class BatchViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = BatchSerializer
    queryset = Batch.objects.all()
    permission_classes = [permissions.IsAuthenticated, IsStaff]

    @detail_route(methods=['get'], permission_classes=[permissions.IsAuthenticated, IsStaff])
    def interns(self, request, pk, *args, **kwargs):
        batch = get_object_or_404(Batch, id=pk)
        interns = Intern.objects.filter(batch=batch).prefetch_related('profile__user', 'internship')
        serialized = InternTableSerializer(interns, many=True)
        return Response(serialized.data)

    @detail_route(methods=['get'], permission_classes=[permissions.IsAuthenticated, IsStaff])
    def plans(self, request, pk, *args, **kwargs):
        batch = get_object_or_404(Batch, id=pk)
        plans = Internship.objects.filter(intern__batch=batch).prefetch_related(
            'rotation_requests__requested_department__department__hospital',
            'rotation_requests__requested_department__department__specialty',
            'rotation_requests__response',
            'rotation_requests__forward',
            'rotations__department__specialty',
            'rotations__department__hospital',
            'intern__profile__user__freezes',
            'intern__profile__user__freeze_requests__response',
            'intern__profile__user__freeze_cancel_requests__response',
            'intern__profile__user__leaves',
            'intern__profile__user__leave_requests__response',
            'intern__profile__user__leave_cancel_requests__response',
            'intern__university',
            'intern__batch',
        ).order_by(
            Lower('intern__profile__en_first_name'),
            Lower('intern__profile__en_father_name'),
            Lower('intern__profile__en_grandfather_name'),
            Lower('intern__profile__en_last_name'),
        )

        filtered = InternshipFilter({'intern__profile__en_full_name': request.query_params.get('query')}, plans)

        paginator = PageNumberPagination()
        paginator.page_size = 10
        page = paginator.paginate_queryset(filtered.qs, request)

        if page is not None:
            serialized = FullInternshipSerializer2(page, many=True)
            return paginator.get_paginated_response(serialized.data)

        serialized = FullInternshipSerializer2(filtered.qs, many=True)
        return Response(serialized.data)

    @detail_route(methods=['get'], permission_classes=[permissions.IsAuthenticated, IsStaff])
    def monthly_list(self, request, pk, *args, **kwargs):
        """
        Return the list of rotations for a given month and department
        """
        batch = get_object_or_404(Batch, id=pk)

        if len(request.query_params.keys()) == 0:
            raise ParseError(detail="No query parameters were specified.")  # FIXME: Is this the most accurate error?

        department = request.query_params.get('department')
        month = request.query_params.get('month')

        if department is None or month is None:
            raise ParseError(detail="Both `department` and `month` query parameters should be specified.")

        month = Month.from_int(int(month))
        department = Department.objects.get(id=department)
        rotations = Rotation.objects.filter(
            internship__intern__batch=batch,
            department=department,
            month=month,
        ).prefetch_related(
            'internship__intern__profile',
        ).order_by(
            Lower('internship__intern__profile__en_first_name'),
            Lower('internship__intern__profile__en_father_name'),
            Lower('internship__intern__profile__en_grandfather_name'),
            Lower('internship__intern__profile__en_last_name'),
        )

        if request.query_params.get('excel'):
            output = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8'
            )
            file_name = "Intern List.xlsx"
            output['Content-Disposition'] = 'attachment; filename=' + file_name

            wb = openpyxl.Workbook()
            ws = wb.active

            title_range = "B2:F2"
            title_cell = "B2"
            subtitle_range = "B3:F3"
            subtitle_cell = "B3"
            stamp_cell = "F4"

            header_range = "B5:F5"

            data_start_row = 6
            data_end_row = data_start_row + rotations.count() - 1
            data_range = "B{}:F{}".format(data_start_row, data_end_row)

            full_range = "B2:F{}".format(data_end_row)
            extended_range = "A1:H{}".format(data_end_row + 2)

            ws.append([])
            ws.append(['', 'Intern List'])
            ws.append(['', 'Month: {} | Department: {} | Batch: {}'.format(
                month.first_day().strftime("%B %Y"),
                department.name,
                batch.name,
            )])
            ws.append(['' for _ in range(5)] + ['Generated by Easy Internship'])

            ws.append(['', "Name", "Student Number", "Badge Number", "Email", "Mobile Number"])
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 35
            ws.column_dimensions['F'].width = 25

            for rotation in rotations:
                ws.append(['',
                    rotation.internship.intern.profile.get_en_full_name(),
                    rotation.internship.intern.student_number,
                    rotation.internship.intern.badge_number,
                    rotation.internship.intern.profile.user.email,
                    rotation.internship.intern.mobile_number,
                ])

            ws.merge_cells(title_range)
            ws.merge_cells(subtitle_range)

            black_border = styles.Border(
                left=styles.Side(border_style='thin', color="000000"),
                top=styles.Side(border_style='thin', color="000000"),
                right=styles.Side(border_style='thin', color="000000"),
                bottom=styles.Side(border_style='thin', color="000000"),
            )
            white_border = styles.Border(
                left=styles.Side(border_style='thin', color="FFFFFF"),
                top=styles.Side(border_style='thin', color="FFFFFF"),
                right=styles.Side(border_style='thin', color="FFFFFF"),
                bottom=styles.Side(border_style='thin', color="FFFFFF"),
            )

            extended_cells = ws[extended_range]
            for x, row in enumerate(extended_cells):
                for y, cell in enumerate(row):
                    # if cell == row[-1]:
                    #     cell.border = styles.Border(
                    #         left=styles.Side(border_style='thin', color='FFFFFF'),
                    #         top=styles.Side(border_style='thin', color='FFFFFF'),
                    #         bottom=styles.Side(border_style='thin', color='FFFFFF'),
                    #     )
                    # elif row == extended_cells[-1]:
                    #     cell.border = styles.Border(
                    #         left=styles.Side(border_style='thin', color='FFFFFF'),
                    #         top=styles.Side(border_style='thin', color='FFFFFF'),
                    #         right=styles.Side(border_style='thin', color='FFFFFF'),
                    #     )
                    # else:
                    cell.border = white_border
                    if cell == row[-1]:
                        cell.border = styles.Border(
                            right=styles.Side(border_style=None),
                        )
                    if row == extended_cells[-1]:
                        cell.border = styles.Border(
                            bottom=styles.Side(border_style=None),
                        )

            def style_range(ws, cell_range, alignment=None, font=None, fill=None, border=None):
                """
                :param ws:  Excel worksheet instance
                :param range: An excel range to style (e.g. A1:F20)
                :param alignment: An openpyxl Alignment object
                :param font: An openpyxl Font object
                :param fill: An openpyxl Fill object
                :param border: An openpyxl Border object
                """

                start_cell, end_cell = cell_range.split(':')
                start_coord = coordinate_from_string(start_cell)
                start_row = start_coord[1]
                start_col = column_index_from_string(start_coord[0])
                end_coord = coordinate_from_string(end_cell)
                end_row = end_coord[1]
                end_col = column_index_from_string(end_coord[0])

                for row in range(start_row, end_row + 1):
                    for col_idx in range(start_col, end_col + 1):
                        col = get_column_letter(col_idx)
                        if alignment:
                            ws.cell('%s%s' % (col, row)).alignment = alignment
                        if font:
                            ws.cell('%s%s' % (col, row)).font = font
                        if fill:
                            ws.cell('%s%s' % (col, row)).fill = fill
                        if border:
                            ws.cell('%s%s' % (col, row)).border = border

            style_range(
                ws, title_range,
                alignment=styles.Alignment(horizontal='center', vertical='center'),
                font=styles.Font(b=True, size=24, color="FFFFFF"),
                fill=styles.PatternFill("solid", fgColor="214a7b"),
                border=black_border,
            )
            style_range(
                ws, subtitle_range,
                alignment=styles.Alignment(horizontal='center', vertical='center'),
                font=styles.Font(i=True, size=14),
                border=black_border,
            )

            stamp_cell = ws[stamp_cell]
            stamp_cell.font = styles.Font(i=True, u='single', color='0000FF')
            stamp_cell.alignment = styles.Alignment(horizontal='right')
            stamp_cell.hyperlink = "https://easyinternship.net"

            header_cells = ws[header_range]
            for row in header_cells:
                for cell in row:
                    cell.font = styles.Font(b=True, size=16)
                    cell.fill = styles.PatternFill('solid', fgColor='ddd9c5')
                    cell.alignment = styles.Alignment(horizontal='center', vertical='center')
                    cell.border = black_border

            data_cells = ws[data_range]
            for row in data_cells:
                for cell in row:
                    cell.font = styles.Font(size=16)
                    cell.border = black_border

            wb.save(output)
            return output

        serialized = FullRotationSerializer(rotations, many=True)
        return Response(serialized.data)
